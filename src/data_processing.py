"""Data loading and cleaning for RMS indentation validation.

The functions in this module intentionally keep a traceable connection to the
source workbook: every cleaned record carries its original sheet name and Excel
row number. The source workbook is read-only input and is never modified.
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook


WORKBOOK_NAME = "Hardness_measurement_results.xlsx"
IMAGE_SUFFIXES = {".bmp", ".png", ".tif", ".tiff"}


@dataclass(frozen=True)
class RepositoryPaths:
    """Resolved repository paths used by the analysis workflow."""

    root: Path
    data: Path
    docs: Path
    images: Path
    notebooks: Path
    outputs: Path


def find_repository_root(start: str | Path | None = None) -> Path:
    """Find the repository root from any child directory.

    The root is recognized by the presence of
    ``data/Hardness_measurement_results.xlsx``. This makes the notebook runnable
    from either the repository root or the ``notebooks`` directory.
    """

    current = Path(start or Path.cwd()).resolve()
    candidates = [current, *current.parents]
    for candidate in candidates:
        if (candidate / "data" / WORKBOOK_NAME).exists():
            return candidate
        if candidate.name == "data" and (candidate / WORKBOOK_NAME).exists():
            return candidate.parent
    raise FileNotFoundError(
        f"Could not find repository root containing data/{WORKBOOK_NAME}"
    )


def get_repository_paths(start: str | Path | None = None) -> RepositoryPaths:
    """Return standard repository paths."""

    root = find_repository_root(start)
    return RepositoryPaths(
        root=root,
        data=root / "data",
        docs=root / "docs",
        images=root / "images",
        notebooks=root / "notebooks",
        outputs=root / "outputs",
    )


def normalize_text(value: Any) -> str:
    """Normalize text labels without changing scientific meaning."""

    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_magnification(value: Any) -> str:
    """Normalize magnification labels such as 20X, 20x and 20 x."""

    text = normalize_text(value).replace(" ", "")
    match = re.fullmatch(r"(\d+(?:[.,]\d+)?)[xX]", text)
    if match:
        number = match.group(1).replace(",", ".")
        if float(number).is_integer():
            number = str(int(float(number)))
        return f"{number}X"
    return text.upper()


def to_float(value: Any) -> float:
    """Convert workbook values to float, accepting decimal commas."""

    if value is None or value == "":
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = normalize_text(value).replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return np.nan


def normalize_scale(method: str, scale_raw: Any, force_n: Any) -> str:
    """Create a normalized hardness-scale label for grouping.

    The original workbook label is retained separately. For Micro-Vickers, the
    workbook contains ``HV 0,5`` labels for both 0.4905 N and 4.905 N rows. The
    normalized scale is therefore derived from force where possible.
    """

    scale = normalize_text(scale_raw).replace(",", ".")
    force = to_float(force_n)
    if method == "Micro-Vickers" and np.isfinite(force):
        kgf = force / 9.81
        if kgf < 0.075:
            return "HV 0.05"
        if kgf < 0.75:
            return "HV 0.5"
        if kgf < 1.5:
            return "HV 1"
        return "HV 2"
    return scale


def hardness_nominal_from_scale(scale_raw: Any, force_n: Any | None = None) -> str:
    """Extract a compact nominal hardness level from a scale/force label."""

    text = normalize_text(scale_raw)
    numbers = re.findall(r"\d+(?:[.,]\d+)?", text)
    if numbers:
        return numbers[0].replace(",", ".")
    if force_n is not None and np.isfinite(to_float(force_n)):
        return f"{to_float(force_n):g} N"
    return ""


def list_image_files(images_root: str | Path) -> pd.DataFrame:
    """List indentation and calibration image files with normalized stems."""

    root = Path(images_root)
    records: list[dict[str, Any]] = []
    for path in sorted(root.rglob("*")):
        if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES:
            parts_lower = {part.lower() for part in path.parts}
            kind = "calibration" if "calibration" in parts_lower else "indentation"
            records.append(
                {
                    "image_kind": kind,
                    "path": path,
                    "relative_path": str(path.relative_to(root.parent))
                    if root.parent in path.parents
                    else path.name,
                    "filename": path.name,
                    "stem": path.stem,
                    "stem_lower": path.stem.lower(),
                    "suffix": path.suffix.lower(),
                }
            )
    return pd.DataFrame(records)


VICKERS_COLUMNS = {
    "serial_number": 1,
    "hardness_scale_raw": 2,
    "position": 3,
    "applied_force_n": 4,
    "magnification_raw": 5,
    "manual_d1_um": 6,
    "manual_d2_um": 7,
    "manual_mean_dimension_um": 8,
    "manual_series_mean_dimension_um": 9,
    "manual_hardness": 10,
    "manual_series_mean_hardness": 11,
    "manual_expanded_uncertainty": 12,
    "manual_relative_expanded_uncertainty": 13,
    "automatic_d1_um": 14,
    "automatic_d2_um": 15,
    "automatic_mean_dimension_um": 16,
    "automatic_series_mean_dimension_um": 17,
    "automatic_hardness": 18,
    "automatic_series_mean_hardness": 19,
    "automatic_expanded_uncertainty": 20,
    "automatic_relative_expanded_uncertainty": 21,
    "picture_id": 22,
}

BRINELL_COLUMNS = {
    "serial_number": 1,
    "hardness_scale_raw": 2,
    "position": 3,
    "ball_diameter_mm": 4,
    "applied_force_n": 5,
    "magnification_raw": 6,
    "manual_d1_um": 7,
    "manual_d2_um": 8,
    "manual_mean_dimension_um": 9,
    "manual_series_mean_dimension_um": 10,
    "manual_hardness": 11,
    "manual_series_mean_hardness": 12,
    "manual_expanded_uncertainty": 13,
    "manual_relative_expanded_uncertainty": 14,
    "automatic_d1_um": 15,
    "automatic_d2_um": 16,
    "automatic_mean_dimension_um": 17,
    "automatic_series_mean_dimension_um": 18,
    "automatic_hardness": 19,
    "automatic_series_mean_hardness": 20,
    "automatic_expanded_uncertainty": 21,
    "automatic_relative_expanded_uncertainty": 22,
    "picture_id": 23,
}

KNOOP_COLUMNS = {
    "serial_number": 1,
    "hardness_scale_raw": 2,
    "position": 3,
    "applied_force_n": 4,
    "magnification_raw": 5,
    "manual_d_um": 6,
    "manual_hardness": 7,
    "automatic_d_um": 8,
    "automatic_hardness": 9,
    "delta_um": 10,
    "delta_percent": 11,
}


def _row_is_measurement(row_values: Iterable[Any]) -> bool:
    values = list(row_values)
    if not any(v not in (None, "") for v in values):
        return False
    serial = normalize_text(values[0] if values else "")
    scale = normalize_text(values[1] if len(values) > 1 else "")
    position = values[2] if len(values) > 2 else None
    if not serial or not scale or position in (None, ""):
        return False
    if serial.lower() == "serial number":
        return False
    return np.isfinite(to_float(position))


def make_series_id(row: pd.Series) -> str:
    """Build the unique three-indentation series identifier."""

    ball = row.get("ball_diameter_mm", np.nan)
    ball_part = "" if pd.isna(ball) else f"|ball={ball:g}mm"
    force = row.get("applied_force_n", np.nan)
    force_part = "" if pd.isna(force) else f"|F={force:g}N"
    return (
        f"{row.get('method')}|SN={row.get('serial_number')}"
        f"|scale={row.get('hardness_scale')}{force_part}{ball_part}"
        f"|mag={row.get('magnification')}"
    )


def _read_sheet(ws, method: str, column_map: dict[str, int]) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    for row_idx in range(3, ws.max_row + 1):
        raw_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if not _row_is_measurement(raw_values):
            continue
        record: dict[str, Any] = {
            "method": method,
            "original_sheet": ws.title,
            "excel_row": row_idx,
        }
        for field, col in column_map.items():
            record[field] = ws.cell(row_idx, col).value
        records.append(record)

    df = pd.DataFrame(records)
    if df.empty:
        return df

    text_cols = ["serial_number", "hardness_scale_raw", "magnification_raw", "picture_id"]
    for col in text_cols:
        if col in df.columns:
            df[col] = df[col].map(lambda v: normalize_text(v) if not pd.isna(v) and v != "" else pd.NA)

    numeric_candidates = [
        c
        for c in df.columns
        if c
        not in {
            "method",
            "original_sheet",
            "excel_row",
            "serial_number",
            "hardness_scale_raw",
            "magnification_raw",
            "picture_id",
        }
    ]
    for col in numeric_candidates:
        df[col] = df[col].map(to_float)

    df["magnification"] = df["magnification_raw"].map(normalize_magnification)
    df["hardness_scale"] = [
        normalize_scale(method, scale, force)
        for scale, force in zip(df["hardness_scale_raw"], df.get("applied_force_n", pd.Series(dtype=float)))
    ]
    df["hardness_level"] = [
        hardness_nominal_from_scale(scale, force)
        for scale, force in zip(df["hardness_scale"], df.get("applied_force_n", pd.Series(dtype=float)))
    ]
    if "ball_diameter_mm" not in df.columns:
        df["ball_diameter_mm"] = np.nan
    df["series_id"] = df.apply(make_series_id, axis=1)
    df["record_id"] = (
        df["method"].astype(str)
        + "|"
        + df["serial_number"].astype(str)
        + "|"
        + df["hardness_scale"].astype(str)
        + "|"
        + df["position"].astype("Int64").astype(str)
        + "|row"
        + df["excel_row"].astype(str)
    )
    return df


def read_hardness_workbook(workbook_path: str | Path) -> dict[str, pd.DataFrame]:
    """Read all recognized workbook sheets into cleaned DataFrames."""

    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True, read_only=False)
    frames: dict[str, pd.DataFrame] = {}
    for sheet in wb.sheetnames:
        lowered = sheet.lower()
        if lowered == "vickers":
            frames["Vickers"] = _read_sheet(wb[sheet], "Vickers", VICKERS_COLUMNS)
        elif lowered in {"mirco-vickers", "micro-vickers"}:
            frames["Micro-Vickers"] = _read_sheet(wb[sheet], "Micro-Vickers", VICKERS_COLUMNS)
        elif lowered == "brinell":
            frames["Brinell"] = _read_sheet(wb[sheet], "Brinell", BRINELL_COLUMNS)
        elif lowered == "knoop":
            frames["Knoop"] = _read_sheet(wb[sheet], "Knoop", KNOOP_COLUMNS)
    return frames


def combine_measurement_frames(frames: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Combine available measurement frames."""

    non_empty = [df for df in frames.values() if df is not None and not df.empty]
    if not non_empty:
        return pd.DataFrame()
    return pd.concat(non_empty, ignore_index=True, sort=False)


def create_series_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Create one row per measurement series from cleaned observations."""

    if df.empty:
        return pd.DataFrame()
    grouped = df.groupby("series_id", dropna=False)
    records: list[dict[str, Any]] = []
    for series_id, group in grouped:
        first = group.iloc[0]
        record = {
            "series_id": series_id,
            "method": first["method"],
            "serial_number": first["serial_number"],
            "hardness_scale_raw": first["hardness_scale_raw"],
            "hardness_scale": first["hardness_scale"],
            "hardness_level": first["hardness_level"],
            "applied_force_n": first.get("applied_force_n", np.nan),
            "ball_diameter_mm": first.get("ball_diameter_mm", np.nan),
            "magnification": first["magnification"],
            "n_positions": int(group["position"].nunique()),
            "excel_rows": ",".join(str(int(r)) for r in group["excel_row"]),
            "picture_ids": "; ".join(
                sorted(
                    {
                        str(v)
                        for v in group.get("picture_id", pd.Series(dtype=object)).dropna()
                        if str(v).strip()
                    }
                )
            ),
        }
        for prefix in ["manual", "automatic"]:
            dim_col = f"{prefix}_mean_dimension_um"
            hard_col = f"{prefix}_hardness"
            record[f"{prefix}_dimension_mean_from_individual_um"] = group[dim_col].mean() if dim_col in group else np.nan
            record[f"{prefix}_dimension_sd_from_individual_um"] = group[dim_col].std(ddof=1) if dim_col in group else np.nan
            record[f"{prefix}_hardness_mean_from_individual"] = group[hard_col].mean() if hard_col in group else np.nan
            record[f"{prefix}_hardness_sd_from_individual"] = group[hard_col].std(ddof=1) if hard_col in group else np.nan
            for field in [
                "series_mean_dimension_um",
                "series_mean_hardness",
                "expanded_uncertainty",
                "relative_expanded_uncertainty",
            ]:
                col = f"{prefix}_{field}"
                record[col] = group[col].dropna().iloc[0] if col in group and group[col].notna().any() else np.nan
        records.append(record)
    return pd.DataFrame(records)


def match_picture_ids(measurements: pd.DataFrame, image_files: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Match workbook Picture IDs to image files by case-insensitive stem."""

    indentation_images = image_files[image_files["image_kind"] == "indentation"].copy()
    image_groups = indentation_images.groupby("stem_lower")
    image_lookup = {stem: group["path"].tolist() for stem, group in image_groups}

    matched_records: list[dict[str, Any]] = []
    for _, row in measurements.iterrows():
        picture_id = row.get("picture_id", pd.NA)
        if pd.isna(picture_id) or not str(picture_id).strip():
            continue
        key = str(picture_id).strip().lower()
        paths = image_lookup.get(key, [])
        status = "matched" if len(paths) == 1 else "missing" if not paths else "ambiguous"
        matched_records.append(
            {
                "method": row["method"],
                "series_id": row["series_id"],
                "record_id": row["record_id"],
                "original_sheet": row["original_sheet"],
                "excel_row": row["excel_row"],
                "picture_id": str(picture_id).strip(),
                "match_status": status,
                "image_path": paths[0] if len(paths) == 1 else pd.NA,
                "n_candidate_files": len(paths),
                "candidate_files": "; ".join(str(p) for p in paths),
            }
        )

    matched = pd.DataFrame(matched_records)
    picture_set = set(matched["picture_id"].str.lower()) if not matched.empty else set()
    unmatched_images = indentation_images[~indentation_images["stem_lower"].isin(picture_set)].copy()
    return matched, unmatched_images


def build_data_quality_report(
    frames: dict[str, pd.DataFrame],
    image_files: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build data-quality and exclusion/reason tables."""

    combined = combine_measurement_frames(frames)
    matched, unmatched_images = match_picture_ids(combined, image_files)
    records: list[dict[str, Any]] = []
    exclusions: list[dict[str, Any]] = []

    for method, df in frames.items():
        if df.empty:
            records.append({"category": "workbook", "method": method, "item": "valid_observations", "value": 0})
            continue
        records.append({"category": "workbook", "method": method, "item": "valid_observations", "value": int(len(df))})
        series_sizes = df.groupby("series_id")["position"].nunique()
        records.append({"category": "workbook", "method": method, "item": "series_count", "value": int(series_sizes.size)})
        records.append({"category": "workbook", "method": method, "item": "complete_three_position_series", "value": int((series_sizes == 3).sum())})
        incomplete = series_sizes[series_sizes != 3]
        records.append({"category": "workbook", "method": method, "item": "incomplete_series", "value": int(len(incomplete))})
        for series_id, count in incomplete.items():
            exclusions.append(
                {
                    "method": method,
                    "scope": "series",
                    "identifier": series_id,
                    "reason": f"Series has {count} unique positions instead of 3",
                }
            )
        for mag, count in Counter(df["magnification"]).items():
            records.append({"category": "workbook", "method": method, "item": f"magnification_{mag}", "value": int(count)})
        pic_count = int(df.get("picture_id", pd.Series(dtype=object)).notna().sum()) if "picture_id" in df else 0
        records.append({"category": "images", "method": method, "item": "picture_ids_in_workbook", "value": pic_count})

        expected_numeric = [
            c
            for c in df.columns
            if c.endswith("_um") or c.endswith("_n") or c.endswith("_mm") or "hardness" in c or "uncertainty" in c
        ]
        for col in expected_numeric:
            records.append(
                {
                    "category": "missing_values",
                    "method": method,
                    "item": col,
                    "value": int(df[col].isna().sum()),
                }
            )

    if not combined.empty and "picture_id" in combined:
        picture_ids = [str(v).strip().lower() for v in combined["picture_id"].dropna() if str(v).strip()]
        duplicate_ids = {k: v for k, v in Counter(picture_ids).items() if v > 1}
        records.append({"category": "images", "method": "All", "item": "duplicate_picture_ids", "value": len(duplicate_ids)})
        for key, count in duplicate_ids.items():
            exclusions.append({"method": "All", "scope": "picture_id", "identifier": key, "reason": f"Duplicate Picture ID appears {count} times"})

    for _, row in matched.iterrows():
        if row["match_status"] != "matched":
            exclusions.append(
                {
                    "method": row["method"],
                    "scope": "picture_id",
                    "identifier": row["picture_id"],
                    "reason": f"Picture ID match status: {row['match_status']}",
                }
            )
    for _, row in unmatched_images.iterrows():
        exclusions.append(
            {
                "method": "Image file",
                "scope": "image",
                "identifier": row["filename"],
                "reason": "Image file has no exact workbook Picture ID match",
            }
        )

    cal_images = image_files[image_files["image_kind"] == "calibration"]
    for _, row in cal_images.iterrows():
        records.append({"category": "calibration", "method": "All", "item": "calibration_image", "value": row["filename"]})

    return pd.DataFrame(records), pd.DataFrame(exclusions)


def rows_for_paired_comparison(df: pd.DataFrame, method: str) -> pd.DataFrame:
    """Return rows with paired manual and automatic data for one method."""

    subset = df[df["method"] == method].copy()
    required = ["manual_mean_dimension_um", "automatic_mean_dimension_um", "manual_hardness", "automatic_hardness"]
    for col in required:
        if col not in subset:
            return pd.DataFrame()
    return subset.dropna(subset=required)


def rows_for_en(series_summary: pd.DataFrame) -> pd.DataFrame:
    """Return series rows suitable for normalized error calculation."""

    required = [
        "manual_series_mean_hardness",
        "automatic_series_mean_hardness",
        "manual_expanded_uncertainty",
        "automatic_expanded_uncertainty",
    ]
    subset = series_summary[series_summary["method"].isin(["Vickers", "Micro-Vickers", "Brinell"])].copy()
    return subset.dropna(subset=required)
